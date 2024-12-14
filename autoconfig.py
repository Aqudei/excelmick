import argparse
import re
import openpyxl
import openpyxl.worksheet

try:
    # from yaml import CLoader as Loader, CDumper as Dumper
    import yaml
    from yaml import CLoader as Loader, CDumper as Dumper
except ImportError:
    # from yaml import Loader, Dumper
    from yaml import Loader, Dumper


def locate_column(column_names, text: str):
    text_normalized = re.sub(r"\s+", "", text).lower()
    for idx, colname in enumerate(column_names):
        colname_normalize = re.sub(r"\s+", "", colname).lower()
        if colname_normalize == text_normalized:
            return idx

    raise Exception(f"A column with name '{text}' was not found in {column_names}")


def qbcc_default(sheet_name: str, column_names):
    sheet_name_lower = sheet_name.lower()

    test_individual = "qbcc" in sheet_name_lower and "individual" in sheet_name_lower
    test_company = "qbcc" in sheet_name_lower and "company" in sheet_name_lower
    test_certifier = "qbcc" in sheet_name_lower and "certifier" in sheet_name_lower
    test_arch = "architects" in sheet_name_lower
    test_engr = "engineers" in sheet_name_lower

    if test_individual or test_company or test_certifier or test_arch or test_engr:
        sheet_config = {
            "license_index": locate_column(column_names, "licence number"),
            "status_index": locate_column(column_names, "status"),
            "last_checked_index": locate_column(column_names, "date checked"),
        }

        return sheet_config


def surveyor(sheet_name: str, column_names):
    sheet_name_lower = sheet_name.lower()

    if "surveyor" in sheet_name_lower:
        sheet_config = {
            "first_name_index": locate_column(column_names, "first name"),
            "surname_index": locate_column(column_names, "surname"),
            "company_index": locate_column(column_names, "company"),
            "status_index": locate_column(column_names, "status"),
            "last_checked_index": locate_column(column_names, "date last checked"),
        }

        return sheet_config


def qbcc_pool_safety(sheet_name: str, column_names):
    sheet_name_lower = sheet_name.lower()

    if "qbcc" in sheet_name_lower and "pool safety" in sheet_name_lower:
        sheet_config = {
            "license_index": locate_column(column_names, "licence number"),
            "status_index": locate_column(column_names, "status"),
            "last_checked_index": locate_column(column_names, "date last checked"),
        }

        return sheet_config


sheet_processors = [qbcc_default, qbcc_pool_safety, surveyor]

def extract_columns(worksheet):
    rows = worksheet.iter_rows()
    row = next(rows)
    if not row:
        raise Exception(f"Workheet '{worksheet.name}' has no available rows")

    while not row[0].value:
        row = next(rows)

    return [c.value for c in row]


def read_config(ymlfile):
    with open(ymlfile) as infile:
        config = yaml.load(infile, Loader=Loader)
        return config


def main():
    """
    docstring
    """
    parser = argparse.ArgumentParser()
    parser.add_argument("input")
    parser.add_argument("--config", default="./config.yml")

    args = parser.parse_args()

    config = read_config(args.config)
    print(config)

    workbook = openpyxl.load_workbook(args.input, read_only=True)
    new_sheet_configs = dict()
    for sheetname in workbook.sheetnames:
        if "how to" in sheetname.lower():
            continue

        for sp in sheet_processors:
            sheet_config = sp(sheetname, extract_columns(workbook[sheetname]))
            if sheet_config:
                new_sheet_configs[sheetname] = sheet_config
                break
    
    config['sheets_config'] = new_sheet_configs
    
    yaml_dumped = yaml.dump(config,Dumper=Dumper)

    # for worksheet in workbook:
    #     columns = extract_columns(worksheet)
    #     print(columns)


if __name__ == "__main__":
    main()
