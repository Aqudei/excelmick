"""
Competent Person checker
"""

import argparse
import time
import json
import logging
import re
import os
import urllib.parse
from functools import partial
from datetime import datetime
import shutil
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver import ChromeOptions
import openpyxl
import requests
import numpy as np
from seleniumrequests import Chrome
import yaml
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

try:
    # from yaml import CLoader as Loader, CDumper as Dumper
    from yaml import CLoader as Loader
except ImportError:
    # from yaml import Loader, Dumper
    from yaml import Loader


# Create a custom logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)  # Set the overall logging level

# Create handlers
console_handler = logging.StreamHandler()  # Handler for console output
file_handler = logging.FileHandler("debug.log")  # Handler for file output

# Set logging levels for handlers
console_handler.setLevel(logging.INFO)  # Only INFO and above for console
file_handler.setLevel(logging.DEBUG)  # DEBUG and above for file

# Create formatters and add it to handlers
formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
console_handler.setFormatter(formatter)
file_handler.setFormatter(formatter)

# Add handlers to the logger
logger.addHandler(console_handler)
logger.addHandler(file_handler)

session = requests.Session()
session.headers.update(
    {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36",  # The version of requests will vary
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
        "Referer": "https://www.onlineservices.qbcc.qld.gov.au/OnlineLicenceSearch/VisualElements/SearchBSALicenseeContent.aspx",
        "Accept-Language": "en-US,en;q=0.8",
        "Cache-Control": "no-cache",
        "Connection": "keep-alive",
    }
)


def enum_rows(sheet):
    """Enumerate Rows

    Args:
        sheet (_type_): _description_

    Yields:
        _type_: _description_
    """

    headers = list()

    for r in sheet.rows:
        values = [f"{c.value}".strip() for c in r]

        if not headers:
            headers = [v.lower() for v in values]
            continue

        item = dict()
        for h, c in zip(headers, r):
            item[h] = f"{c.value}".strip() if c else ""

        yield r, item


def parse_qbcc_response(html):
    soup = BeautifulSoup(html, "html.parser")

    business_name_element = soup.select_one(
        "#ctl00_generalContentPlaceHolder_LicenceInfoControl1_lbLicenceName"
    )
    trading_name_element = soup.select_one(
        "ctl00_generalContentPlaceHolder_LicenceInfoControl1_lbTradingName"
    )

    items = [
        td.text.strip("\r\n\t ")
        for td in soup.select(
            "table[id='ctl00_generalContentPlaceHolder_LicenceInfoControl1_gvLicenceClass'] td"
        )
    ]

    items_array = np.array(items)

    if len(items_array) > 0 and (len(items_array) % 4) == 0:
        for part in np.split(items_array, len(items_array) / 4):
            yield [p.strip() for p in part]


def parse_surveyor_response(html_content):
    """
    parse surveyor response
    """
    # Parse the HTML using BeautifulSoup
    soup = BeautifulSoup(html_content, "html.parser")
    element = soup.select_one(".search-results")
    if not element:
        return

    # Extract name
    name = element.find("h4").get_text(strip=False).split("<br>")[0]
    lic_class = element.find("h4").get_text(strip=False).split("<br>")[-1]
    # Extract phone number
    phone_element = element.find("td", string="Phone ")
    phone = ""
    if phone_element:
        phone = phone_element.find_next_sibling("td").get_text()

    # Extract email
    email_element = element.find("td", string="Email ")
    email = ""
    if email_element:
        email = email_element.find_next_sibling("td").get_text()

    # Extract address
    address_element = element.find("td", string="Address")
    address = ""
    if address_element:
        address = address_element.find_next_sibling("td").get_text(separator=", ")

    types = []
    # Extract types
    types = [
        span.get_text()
        for span in element.find_all("div", class_="types")[0].find_all("span")
    ]

    return {
        "name": re.sub(r"\s+", " ", name),
        "license_class": re.sub(r"\s+", " ", lic_class),
        "phone": phone,
        "email": email,
        "address": address,
        "expertise": "; ".join(types),
    }


def query_surveyor_license(search_text):
    """
    query surveyor license
    """
    search_text = re.sub(r"\s+", " ", f"{search_text}".strip())
    logger.info("Looking up surveyor info: %s", search_text)
    url = "https://sbq.com.au/find-a-surveyor/search-cadastral/"
    session.get(url)

    params = {
        "surveyor-type": "individual",
        "search-type": "name",
        "title": search_text,
        "postcode": "",
        "radius": "5",
        "q": "Search",
    }

    response = session.get(url, params=params)
    if response.status_code != 200:
        return

    return parse_surveyor_response(response.text)


def query_qbcc_license(license_no):
    """
    query qbcc license
    """

    session.get(
        "https://www.onlineservices.qbcc.qld.gov.au/OnlineLicenceSearch/VisualElements/SearchBSALicenseeContent.aspx"
    )

    license_no = f"{license_no}".strip("\r\n\t ")
    url = "http://www.onlineservices.qbcc.qld.gov.au/OnlineLicenceSearch/VisualElements/ShowDetailResultContent.aspx"
    params = {
        "LicNO": f"{license_no}",
        "licCat": "LIC",
        "name": "",
        "firstName": "",
        "searchType": "Contractor",
        "FromPage": "SearchContr",
    }
    response = requests.get(url, params=params)
    for r in parse_qbcc_response(response.text):
        yield r


def query_qbcc_certifier_license(license_no):
    """
    query_qbcc_certifier_license
    """
    session.get(
        "https://www.onlineservices.qbcc.qld.gov.au/OnlineLicenceSearch/VisualElements/SearchBuildingCertifierContent.aspx"
    )

    license_no = license_no.strip("\r\n\t ")
    url = "https://www.onlineservices.qbcc.qld.gov.au/OnlineLicenceSearch/VisualElements/ShowDetailResultContent.aspx"
    params = {
        "LicNO": f"{license_no}",
        "licCat": "LIC",
        "name": "",
        "firstName": "",
        "searchType": "Certifier",
        "FromPage": "SearchContr",
    }

    response = session.get(url, params=params)
    for r in parse_qbcc_response(response.text):
        yield r


def query_engr_registration(license_number, driver: Chrome):
    """
    query_engr_registration
    """
    try:
        url1 = "https://portal.bpeq.qld.gov.au/BPEQPortal/RPEQ_Directory.aspx"
        driver.get(url1)
        element = WebDriverWait(driver, 16).until(
            EC.presence_of_element_located(
                (
                    By.ID,
                    "ctl01_TemplateBody_WebPartManager1_gwpciEngineersearch_ciEngineersearch_ResultsGrid_Sheet0_Input3_TextBox1",
                )
            )
        )
        element.send_keys(license_number)
        element = WebDriverWait(driver, 16).until(
            EC.presence_of_element_located(
                (
                    By.ID,
                    "ctl01_TemplateBody_WebPartManager1_gwpciEngineersearch_ciEngineersearch_ResultsGrid_Sheet0_SubmitButton",
                )
            )
        )
        element.click()
        element = WebDriverWait(driver, 16).until(
            EC.presence_of_element_located(
                (
                    By.XPATH,
                    '//*[@id="ctl01_TemplateBody_WebPartManager1_gwpciEngineersearch_ciEngineersearch_ResultsGrid_Grid1_ctl00__0"]/td[1]/a',
                )
            )
        )
        registration_no = element.get_attribute("href").split("=")[1]

        url = f"https://portal.bpeq.qld.gov.au/Party.aspx?ID={registration_no}"
        response = driver.request("GET", url, verify=False)
        soup = BeautifulSoup(response.text, "html.parser")
        parts = [
            p.text.strip("\r\n\t ") for p in soup.select(".PanelFieldValue > span")
        ]
        logger.info("Num Parts: %d", len(parts))

        name = soup.title.text.strip("\r\n\t")
        date_registered_from = parts[0]
        status = parts[3]
        date_registered_to = parts[4]
        company = parts[5]
        job_type = parts[1]

        return {
            "name": name,
            "company": company,
            "date_registered_from": date_registered_from,
            "job_type": job_type,
            "status": status,
            "date_registered_to": date_registered_to,
        }

    except Exception as e:
        logger.info(e)


def query_arch_registration(license_number, driver: Chrome):
    try:
        url1 = "https://www.boaq.qld.gov.au/Web/Consumers/Search_the_Register/Web/Architect_Search.aspx?hkey=f493b110-1ad9-4ec8-a830-f9a1f70e16b5"
        driver.get(url1)
        element = WebDriverWait(driver, 16).until(
            EC.presence_of_element_located(
                (
                    By.ID,
                    "ctl01_TemplateBody_WebPartManager1_gwpciArchitectsearch_ciArchitectsearch_ResultsGrid_Sheet0_Input3_TextBox1",
                )
            )
        )
        element.send_keys(license_number)
        element = WebDriverWait(driver, 16).until(
            EC.presence_of_element_located(
                (
                    By.NAME,
                    "ctl01$TemplateBody$WebPartManager1$gwpciArchitectsearch$ciArchitectsearch$ResultsGrid$Sheet0$SubmitButton",
                )
            )
        )
        element.click()
        element = WebDriverWait(driver, 16).until(
            EC.presence_of_element_located(
                (
                    By.XPATH,
                    '//*[@id="ctl01_TemplateBody_WebPartManager1_gwpciArchitectsearch_ciArchitectsearch_ResultsGrid_Grid1_ctl00__0"]/td[1]/a',
                )
            )
        )
        registration_no = element.get_attribute("href").split("=")[1]
        url = f"https://www.boaq.qld.gov.au/Party.aspx?ID={registration_no}"
        response = driver.request("GET", url)
        soup = BeautifulSoup(response.text, "html.parser")
        parts = [
            p.text.strip("\r\n\t ") for p in soup.select(".PanelFieldValue > span")
        ]
        logger.info("Num Parts: %d", len(parts))
        if len(parts) == 12:
            name = parts[0]
            company = parts[1]
            job_type = parts[2]
            date_joined = parts[4]
            status = parts[3]
            date_registered = parts[4]

            return name, company, date_joined, job_type, status, date_registered
        elif len(parts) == 11:
            name = parts[0]
            date_joined = parts[3]
            job_type = parts[1]
            status = parts[2]
            date_registered = parts[3]

            return name, None, date_joined, job_type, status, date_registered

    except Exception as e:
        logger.info(e)


def process_sheet_arch(wb, sheetname, args, config, sheet_config, orig_filename):
    """
    process_sheet_arch
    """
    if "architects" not in sheetname.lower():
        return

    logger.info("Processing Architects Tab...")
    sheet = wb[sheetname]

    config = read_config()
    count = 0

    driver = init_web_driver()

    for idx, (row, data) in enumerate(enum_rows(sheet)):
        if should_skip_row(row, sheet_config, config):
            continue

        if count > 0 and (count % config["numrec_before_save"]) == 0:
            logger.info(
                "===============================================\n \
                Saving progress to excel file...\n=================================="
            )
            wb.save(orig_filename)

        registration_no = str(row[sheet_config["license_index"]].value or "")

        logger.info("Fetching Registration info of %s:", registration_no)
        reg_status = query_arch_registration(registration_no, driver)

        if reg_status:
            logger.info("Registration info found!")

            name, company, date_joined, job_type, status, date_registered = reg_status

            # lic_class, _, _, lic_status = lic_statuses[0]
            logger.info("\tName: %s", name)
            logger.info("\tCompany: %s", company)
            logger.info("\tDate Joined: %s", date_joined)
            logger.info("\tType: %s", job_type)
            logger.info("\tStatus: %s", status)
            logger.info("\tDate Registered: %s", date_registered)

            row[sheet_config["status_index"]].value = status.strip().title()
            row[sheet_config["last_checked_index"]].value = datetime.now().date()
        else:
            logger.info("Registration info not found in online register !")
            row[sheet_config["status_index"]].value = "Missing in Register"
            row[sheet_config["last_checked_index"]].value = datetime.now().date()

        count = count + 1


def init_web_driver():
    """
    init_chrome
    """
    options = ChromeOptions()
    # chrome-win64\chrome.exe
    options.binary_location = os.path.join("chrome-win64", "chrome.exe")
    # options.add_argument("headless")
    driver = Chrome(options=options)
    return driver


def process_sheet_engr(wb, sheetname, args, config, sheet_config, orig_filename):
    """
    process_sheet_engr
    """
    if not all(keyword in sheetname.lower() for keyword in ["engineers"]):
        return

    logger.info("Processing Engineers Tab...")
    sheet = wb[sheetname]

    config = read_config()
    count = 0

    driver = init_web_driver()

    for row, data in enum_rows(sheet):
        if count > 0 and (count % config["numrec_before_save"]) == 0:
            logger.info(
                "===============================================\n \
                Saving progress to excel file...\n=================================="
            )
            wb.save(orig_filename)

        if should_skip_row(row, sheet_config, config):
            continue

        license_number = str(row[sheet_config["license_index"]].value or "")
        logger.info("Fetching Registration info of %s:", license_number)
        reg_status = query_engr_registration(license_number, driver)

        if reg_status:
            logger.info("Registration info found!")

            # name, company, date_registered_from, job_type, status, date_registered = reg_status
            # return name, company, date_registered_from, job_type, status, date_registered_to
            # lic_class, _, _, lic_status = lic_statuses[0]

            status = reg_status["status"]

            logger.info("\tName: %s", reg_status["name"])
            logger.info("\tCompany: %s", reg_status["company"])
            logger.info(
                "\tDate Registered From: %s", reg_status["date_registered_from"]
            )
            logger.info("\tType: %s", reg_status["job_type"])
            logger.info("\tStatus: %s", reg_status["status"])
            logger.info("\tDate Registered To: %s", reg_status["date_registered_to"])

            row[sheet_config["status_index"]].value = status.strip().title()
            row[sheet_config["last_checked_index"]].value = datetime.now().date()
        else:
            logger.info("Registration info not found in online register !")
            row[sheet_config["status_index"]].value = "Missing in Register"
            row[sheet_config["last_checked_index"]].value = datetime.now().date()

        count = count + 1


def update_license_status(row, status, sheet_config):
    """Helper function to update status and last checked date for a row."""
    row[sheet_config["status_index"]].value = status
    row[sheet_config["last_checked_index"]].value = datetime.now().date()


def handle_surveyor_license_query(row, search_text, sheet_config):
    """Query surveyor's license and update the status in the row."""
    result = query_surveyor_license(search_text)
    if result:
        update_license_status(row, "Active", sheet_config)
        return True
    else:
        update_license_status(row, "License Not Found", sheet_config)
        return False


def process_sheet_surveyor(wb, sheetname, args, config, sheet_config, orig_filename):
    """
    process_sheet_surveyor
    """
    if "surveyor" not in sheetname.lower():
        return

    logger.info("Processing Surveyor Tab: %s...", sheetname)
    session.headers.clear()
    session.headers.update(
        {
            "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
            "accept-language": "en-US,en;q=0.7",
            "cache-control": "no-cache",
            "pragma": "no-cache",
            "referer": "https://sbq.com.au/find-a-surveyor/search-cadastral/",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36",
        }
    )

    sheet = wb[sheetname]
    count = 0

    for idx, (row, data) in enumerate(enum_rows(sheet)):
        if count > 0 and (count % config["numrec_before_save"]) == 0:
            logger.info(
                "===============================================\n \
                Saving progress to excel file...\n=================================="
            )
            wb.save(orig_filename)

        # Skip rows with a recent "last checked" date
        if should_skip_row(row, sheet_config, config):
            continue

        first_name = f"{row[sheet_config['first_name_index']].value or ''}".strip()
        surname = f"{row[sheet_config['surname_index']].value or ''}".strip()
        search_text = re.sub(r"\s+", " ", f"{first_name} {surname}".strip())

        if first_name == "" or surname == "":
            company = f"{row[sheet_config['company_index']].value or ''}".strip()
            handle_surveyor_license_query(row, company, sheet_config)
        else:
            # First try with name, fallback to company if name fails
            if not handle_surveyor_license_query(row, search_text, sheet_config):
                company = f"{row[sheet_config['company_index']].value or ''}".strip()
                handle_surveyor_license_query(row, company, sheet_config)

        count += 1


def should_skip_row(row, sheet_config, cfg):
    """Check if the row should be skipped based on last checked date."""
    last_date_checked = row[sheet_config["last_checked_index"]].value
    if last_date_checked and isinstance(last_date_checked, datetime):
        delta = datetime.now().date() - last_date_checked.date()
        return delta.days <= cfg.get("skip_days", 5)
    else:
        return False


def query_pool_safety_license(lic_no):
    """
    query_pool_safety_license
    """
    default_headers = {
        "accept": "*/*",
        "accept-language": "en-US,en;q=0.9",
        "cache-control": "no-cache",
        "content-type": "application/x-www-form-urlencoded; charset=UTF-8",
        "origin": "https://my.qbcc.qld.gov.au",
        "pragma": "no-cache",
        "priority": "u=1, i",
        "referer": "https://my.qbcc.qld.gov.au/s/pool-safety-inspector-search",
        "sec-ch-ua": '"Brave";v="129", "Not=A?Brand";v="8", "Chromium";v="129"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "sec-fetch-site": "same-origin",
        "sec-gpc": "1",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36",
        # 'x-sfdc-page-scope-id': '3bdc58c4-042b-42d8-8bcd-b875c8aa3bfb',
        # 'x-sfdc-request-id': '415150000004b0f99f',
    }

    s = requests.Session()
    response = s.get("https://my.qbcc.qld.gov.au/s/pool-safety-inspector-search")

    cookies = s.cookies.get_dict()
    context = cookies.get("renderCtx")
    context_decoded = json.loads(urllib.parse.unquote(context))
    sfdc_req_id = response.headers.get("x-sfdc-request-id")
    x_sfdc_page_scope_id = context_decoded.get("pageId")

    s.headers.update(default_headers)
    s.headers.update(
        {
            # "x-request-id":req_id,
            "X-Sfdc-Request-Id": sfdc_req_id,
            "X-Sfdc-Page-Scope-Id": x_sfdc_page_scope_id,
        }
    )
    url = "https://my.qbcc.qld.gov.au/s/sfsites/aura?other.PSISearch.searchInspectors=1"

    data = {
        "message": '{"actions":[{"id":"175;a","descriptor":"apex://PSISearchController/ACTION$searchInspectors","callingDescriptor":"markup://c:PSI_Search","params":{"searchBy":"licence","firstName":"","lastName":"","businessName":"","licenceNumber":"%s","distanceInKm":5,"batchSize":1000,"offset":0}}]}'
        % (lic_no,),
        "aura.context": '{"mode":"PROD","fwuid":"eGx3MHlRT1lEMUpQaWVxbGRUM1h0Z2hZX25NdHFVdGpDN3BnWlROY1ZGT3cyNTAuOC40LTYuNC41","app":"siteforce:communityApp","loaded":{"APPLICATION@markup://siteforce:communityApp":"wi0I2YUoyrm6Lo80fhxdzA","COMPONENT@markup://instrumentation:o11ySecondaryLoader":"1JitVv-ZC5qlK6HkuofJqQ"},"dn":[],"globals":{},"uad":false}',
        "aura.pageURI": "/s/pool-safety-inspector-search",
        "aura.token": "null",
    }

    response = s.post(url, data=data)

    results = response.json()["actions"][0]["returnValue"]
    if not results or len(results) <= 0:
        return

    results0 = results[0]
    expiry_date = datetime.strptime(results0["expiryDate"], "%Y-%m-%d")
    if datetime.now() > expiry_date:
        results0["expired"] = True

    return results0


def process_sheet_qbcc_pool_safety(
    wb, sheetname, args, config, sheet_config, orig_filename
):
    """
    process_sheet_qbcc_pool_safety
    """
    if not all(keyword in sheetname.lower() for keyword in ["qbcc", "pool", "safety"]):
        return

    logger.info("Processing QBCC Pool Safety Tab <QBCC>...")
    sheet = wb[sheetname]
    save_interval = config["numrec_before_save"]

    for count, (row, data) in enumerate(enum_rows(sheet), start=1):
        logger.info("Processing Line #%d", count)

        if count % save_interval == 0:
            logger.info("Saving progress to excel file...")
            wb.save(orig_filename)

        license_no = data.get("licence number", "").strip()
        # Skip rows with a recent "last checked" date

        if not license_no:
            message = (
                "License No. Column not found!"
                if "licence number" not in data
                else "License No is BLANK !"
            )
            logger.info(message)
            row[sheet_config["status_index"]].value = message
            row[sheet_config["last_checked_index"]].value = datetime.now().date()
            continue

        if should_skip_row(row, sheet_config, config):
            continue

        logger.info("Fetching License info of %s:", license_no)
        lic_status = query_pool_safety_license(license_no)
        if lic_status:
            expired = lic_status.get("expired", False)

            if expired:
                row[sheet_config["status_index"]].value = "License Expired"
            else:
                row[sheet_config["status_index"]].value = "Active"
        else:
            logger.info("License not found in online register!")
            row[sheet_config["status_index"]].value = "Missing in Register"

        row[sheet_config["last_checked_index"]].value = datetime.now().date()

    wb.save(orig_filename)


def process_sheet_qbcc_individual(
    wb,
    sheetname,
    args,
    config,
    sheet_config,
    orig_filename,
    license_querier=query_qbcc_license,
    keywords=None,
):
    """
    process_sheet_qbcc_individual
    """
    used_keywords = []
    if keywords is None:
        used_keywords = ["qbcc", "individual"]
    else:
        used_keywords.extend(keywords)

    if not all(keyword in sheetname.lower() for keyword in used_keywords):
        return

    logger.info("Processing QBCC Individual...")

    sheet = wb[sheetname]
    count = 0

    for idx, (row, data) in enumerate(enum_rows(sheet)):
        logger.info("Processing Line #%s", (idx + 1))

        if should_skip_row(row, sheet_config, config):
            continue

        try_save(wb, config, orig_filename, count)

        license_no = (
            row[sheet_config["license_index"]].value
            if row[sheet_config["license_index"]].value
            else ""
        )
        if license_no in [None, ""]:
            row[sheet_config["status_index"]].value = "Invalid License Number!"
            row[sheet_config["last_checked_index"]].value = datetime.now().date()
            count = count + 1
            continue

        logger.info("Fetching License info of %s:", license_no)
        lic_statuses = list(license_querier(license_no))
        if len(lic_statuses) > 0:
            logger.info("License info found!")
            lic_class, _, _, lic_status = lic_statuses[0]
            logger.info("\tLicense Class: %s", lic_class)
            logger.info("\tStatus: %s", lic_status)

            row[sheet_config["status_index"]].value = lic_status.title().strip()

        else:
            logger.info("License not found in online register !")
            row[sheet_config["status_index"]].value = "Missing in Register"

        row[sheet_config["last_checked_index"]].value = datetime.now().date()
        count = count + 1


def try_save(wb, config, orig_filename, count):
    """
    try saving excel file
    """
    if count > 0 and (count % config["numrec_before_save"]) == 0:
        logger.info(
            "===============================================\n \
                Saving progress to excel file...\n=================================="
        )
        wb.save(orig_filename)


def read_config():
    """
    Read YAML configuration
    """
    with open("./config.yml", "rt", errors="ignore", encoding="utf-8") as fp:
        conf = yaml.load(fp, Loader=Loader)

    return conf


def process_workbook(filepath, args):
    """
    process workbook
    """
    wb = None

    try:
        wb = openpyxl.load_workbook(filepath)

        logger.info("Found %d sheets.", len(wb.sheetnames))
        logger.info("\n".join([f"\t{s}" for s in wb.sheetnames]))

        config = read_config()
        process_qbcc_certifier = partial(
            process_sheet_qbcc_individual,
            license_querier=query_qbcc_certifier_license,
            keywords=["qbcc", "certifier"],
        )
        process_qbcc_company = partial(
            process_sheet_qbcc_individual, keywords=["qbcc", "company"]
        )

        processors = [
            process_sheet_qbcc_individual,
            process_qbcc_company,
            process_qbcc_certifier,
            process_sheet_qbcc_pool_safety,
            process_sheet_surveyor,
            process_sheet_arch,
            process_sheet_engr,
        ]

        if config.get("with_browser", False):
            processors += [
                process_sheet_arch,
                process_sheet_engr,
            ]

        for sheetname in wb.sheetnames:
            for sheetname_filter in config["sheets_config"].keys():
                if sheetname_filter.lower().strip() == sheetname.lower().strip():
                    logger.info("Processing SHEET: %s", sheetname)
                    sheet_config = config["sheets_config"].get(sheetname_filter)
                    if not sheet_config:
                        logger.error(
                            "Sheet/Tab with name: %s was not found in config.yml. Please re-check!",
                            sheetname,
                        )
                        break

                    for processor in processors:
                        processor(wb, sheetname, args, config, sheet_config, filepath)

        logger.info("Process done. Saving workbook to %s.", filepath)
        wb.save(filepath)
    except Exception as e:
        raise e
    finally:
        if wb:
            wb.close()


class IdleFileHandler(FileSystemEventHandler):
    """
    hotfolder watcher class
    """

    def __init__(self, idle_time):
        self.idle_time = idle_time
        self.last_modified_time = {}

    def on_created(self, event):
        if not event.is_directory:
            file_path = event.src_path
            if "~" not in file_path:
                self.last_modified_time[file_path] = time.time()
                logger.info(
                    "New file added: %s, waiting for it to become idle.", file_path
                )

    def on_modified(self, event):
        if not event.is_directory:
            pass

    def __move_file(self, src, dest):
        """
        docstring
        """
        try:
            shutil.move(src, dest)
        except:
            pass

    def process_if_idle(self, file_path, args, config):
        """
        main processor function
        """
        # Wait until the file is idle
        while True:
            if (
                file_path in self.last_modified_time
                and self.last_modified_time[file_path]
            ):
                time_since_modification = (
                    time.time() - self.last_modified_time[file_path]
                )
                if time_since_modification > self.idle_time:
                    print(f"{file_path} is idle, processing...")
                    self.last_modified_time[file_path] = None
                    processing_path = os.path.join(
                        config["processing"], os.path.basename(file_path)
                    )
                    self.__move_file(file_path, processing_path)

                    # Process the file here
                    try:
                        process_workbook(processing_path, args)
                        done_path = os.path.join(
                            config["done"], os.path.basename(file_path)
                        )
                        self.__move_file(processing_path, done_path)
                    except Exception as e:
                        logger.exception(e)
                        error_path = os.path.join(
                            config["error"], os.path.basename(file_path)
                        )
                        self.__move_file(processing_path, error_path)

                    break
            time.sleep(1)


def prep_dirs(config):
    """
    ensure required folders exists
    """
    dirs = [
        config.get("hotfolder"),
        config.get("processing"),
        config.get("done"),
        config.get("error"),
    ]

    for d in dirs:
        if d:
            os.makedirs(d, exist_ok=True)


def main():
    """
    main entry point
    """
    parser = argparse.ArgumentParser()
    args = parser.parse_args()

    config = read_config()

    prep_dirs(config)

    event_handler = IdleFileHandler(config.get("idle_time", 5))
    observer = Observer()
    observer.schedule(event_handler, config["hotfolder"], recursive=False)
    observer.start()

    try:
        logger.info(
            "Monitoring folder: <%s> for changes...",
            os.path.abspath(config["hotfolder"]),
        )
        while True:
            # If a file has been modified, check if it is idle and process it
            for file in event_handler.last_modified_time:
                event_handler.process_if_idle(file, args, config)

            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()

    observer.join()


if __name__ == "__main__":
    main()
